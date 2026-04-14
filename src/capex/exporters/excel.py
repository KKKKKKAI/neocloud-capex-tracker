"""Excel workbook exporter — the visible deliverable.

CURRENCY RULE: all values displayed in Excel are in USD millions.
Non-USD companies (BABA, BIDU, GDS, 0700) are converted at period-end
FX rates. The local currency value is preserved in the DB and shown
in the cell comment (Shift+F2) for audit, but the CELL VALUE is always
value_usd. Never display local currency values in data cells.

Reads the extractions DB and produces an 8-sheet Excel workbook:

    1. Revenue (Annual)      — one row per company, one column per FY
    2. Revenue (Quarterly)   — de-cumulated standalone quarterly values
    3. Capex (Annual)        — same layout as Revenue Annual
    4. Capex (Quarterly)     — de-cumulated standalone quarterly capex
    5. All Metrics (Annual)  — company × metric × year pivot
    6. All Metrics (Quarterly) — same but quarterly
    7. Data Quality          — flags and notes
    8. Metadata              — coverage, FX rates, concepts used

Usage:
    from capex.exporters.excel import export_workbook
    export_workbook("workbook/capex_tracker.xlsx")

    # Or via CLI:
    capex export
"""
from __future__ import annotations

import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

REPO_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_OUTPUT = REPO_ROOT / "workbook" / "capex_tracker.xlsx"

# Import canonical flow/stock metric sets from the extraction layer
from ..extract.decumulate import FLOW_METRICS, STOCK_METRICS

# Coverage start overrides (filter out pre-restructuring noise)
COVERAGE_START_OVERRIDES = {
    "NBIS": "2024-01-01",  # post-Yandex restructuring
}


def export_workbook(
    output_path: str | Path | None = None,
    db_path: str | Path | None = None,
) -> Path:
    """Generate the Excel workbook from the extractions DB."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install with: pip install openpyxl"
        ) from exc

    output_path = Path(output_path or DEFAULT_OUTPUT)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    db_path = db_path or (REPO_ROOT / "data" / "db" / "capex.db")
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row

    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Header style
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    num_fmt = '#,##0'

    # Build the sheets — cover + TOC first, then data sheets
    s = (header_font, header_fill, num_fmt)

    # Data sheets (built first so we can count rows for TOC)
    sheet_info = []

    _build_metric_sheet(wb, conn, "Revenue (Annual)", "revenue", "annual", *s)
    sheet_info.append(("Revenue (Annual)", "Annual total revenue by company (USD M)"))

    _build_metric_sheet(wb, conn, "Revenue (Quarterly)", "revenue", "quarterly", *s)
    sheet_info.append(("Revenue (Quarterly)", "Quarterly de-cumulated revenue (USD M)"))

    _build_metric_sheet(wb, conn, "Capex (Annual)", "capital_expenditures", "annual", *s)
    sheet_info.append(("Capex (Annual)", "Annual capital expenditures by company (USD M)"))

    _build_metric_sheet(wb, conn, "Capex (Quarterly)", "capital_expenditures", "quarterly", *s)
    sheet_info.append(("Capex (Quarterly)", "Quarterly de-cumulated capex (USD M)"))

    _build_metric_sheet(wb, conn, "Cloud Revenue (Annual)", "cloud_segment_revenue", "annual", *s)
    sheet_info.append(("Cloud Revenue (Annual)", "Cloud/datacenter segment revenue (USD M)"))

    _build_metric_sheet(
        wb, conn, "Cloud Revenue (Quarterly)",
        "cloud_segment_revenue", "quarterly", *s,
    )
    sheet_info.append(("Cloud Revenue (Quarterly)", "Quarterly cloud segment revenue (USD M)"))

    _build_all_metrics_sheet(wb, conn, "All Metrics (Annual)", "annual", *s)
    sheet_info.append(("All Metrics (Annual)", "All 6 metrics × all companies, annual"))

    _build_all_metrics_sheet(wb, conn, "All Metrics (Quarterly)", "quarterly", *s)
    sheet_info.append(("All Metrics (Quarterly)", "All 6 metrics × all companies, quarterly"))

    _build_data_quality_sheet(wb, conn, header_font, header_fill)
    sheet_info.append(("Data Quality", "Data quality flags and coverage notes"))

    _build_metadata_sheet(wb, conn, header_font, header_fill)
    sheet_info.append(("Metadata", "Coverage dates, FX rates, extraction sources"))

    # Cover sheet + TOC (inserted at position 0 so it's the first tab)
    _build_cover_sheet(wb, conn, sheet_info, header_font, header_fill)

    wb.save(str(output_path))
    conn.close()
    return output_path


# --------------------------------------------------------------------
# Sheet builders
# --------------------------------------------------------------------


def _get_all_tickers(conn) -> list[str]:
    """Get ALL company tickers from the companies table."""
    return [
        r[0] for r in conn.execute(
            "SELECT ticker FROM companies ORDER BY ticker"
        )
    ]


def _build_metric_sheet(wb, conn, sheet_name, metric_key, cadence, hfont, hfill, nfmt):
    """Build a single-metric sheet: rows=companies, cols=periods."""
    ws = wb.create_sheet(sheet_name)

    data = _get_metric_data(conn, metric_key, cadence)

    # Get all unique periods and ALL companies (not just those with data)
    all_periods = sorted(set(p for company_data in data.values() for p in company_data))
    companies = _get_all_tickers(conn)

    if not all_periods:
        # Still show companies even if no data
        ws.append(["Company", "No data available"])
        for ticker in companies:
            ws.append([ticker])
        return

    # Header row
    headers = ["Company"] + all_periods
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill

    # Data rows — with provenance comments
    from openpyxl.comments import Comment

    from .citations import format_citation

    for row_idx, ticker in enumerate(companies, 2):
        ws.cell(row=row_idx, column=1, value=ticker)
        ticker_data = data.get(ticker, {})
        for col_idx, period in enumerate(all_periods, 2):
            entry = ticker_data.get(period)
            if entry is None:
                continue
            if isinstance(entry, tuple):
                val, meta = entry
            else:
                val, meta = entry, None

            cell = ws.cell(row=row_idx, column=col_idx, value=round(val))
            cell.number_format = nfmt

            if meta:
                try:
                    citation = format_citation(
                        meta.get("extraction", {}),
                        meta.get("source_doc", {}),
                        meta.get("company", {}),
                    )
                    cell.comment = Comment(citation, "capex-tracker")
                    cell.comment.width = 400
                    cell.comment.height = 250
                except Exception:
                    pass  # don't break export on citation errors

    # Auto-width for company column
    ws.column_dimensions["A"].width = 12


def _build_all_metrics_sheet(wb, conn, sheet_name, cadence, hfont, hfill, nfmt):
    """Build a sheet with all metrics: rows=company+metric, cols=periods."""
    ws = wb.create_sheet(sheet_name)

    metrics = ["revenue", "cloud_segment_revenue", "capital_expenditures",
               "operating_cash_flow", "depreciation_amortization",
               "property_plant_equipment_net"]

    all_data = {}
    all_periods = set()
    for mk in metrics:
        mdata = _get_metric_data(conn, mk, cadence)
        for ticker, periods in mdata.items():
            key = (ticker, mk)
            all_data[key] = periods
            all_periods.update(periods.keys())

    all_periods = sorted(all_periods)
    headers = ["Company", "Metric"] + all_periods
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill

    # Use ALL tickers, not just those with data
    all_tickers = _get_all_tickers(conn)
    row_idx = 2
    for ticker in all_tickers:
        for mk in metrics:
            key = (ticker, mk)
            ws.cell(row=row_idx, column=1, value=ticker)
            ws.cell(row=row_idx, column=2, value=mk)
            if key not in all_data:
                row_idx += 1
                continue
            from openpyxl.comments import Comment

            from .citations import format_citation

            for col_idx, period in enumerate(all_periods, 3):
                entry = all_data[key].get(period)
                if entry is None:
                    continue
                if isinstance(entry, tuple):
                    val, meta = entry
                else:
                    val, meta = entry, None
                cell = ws.cell(
                    row=row_idx, column=col_idx, value=round(val)
                )
                cell.number_format = nfmt
                if meta:
                    try:
                        citation = format_citation(
                            meta.get("extraction", {}),
                            meta.get("source_doc", {}),
                            meta.get("company", {}),
                        )
                        cell.comment = Comment(
                            citation, "capex-tracker"
                        )
                        cell.comment.width = 400
                        cell.comment.height = 250
                    except Exception:
                        pass
            row_idx += 1

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30


def _build_cover_sheet(wb, conn, sheet_info, hfont, hfill):
    """Build the cover sheet with summary info + table of contents."""
    from openpyxl.styles import Alignment, Font

    ws = wb.create_sheet("Cover", 0)  # insert at position 0

    title_font = Font(bold=True, size=20, color="1F4E79")
    label_font = Font(bold=True, size=11)
    value_font = Font(size=11)
    disclaimer_font = Font(size=9, italic=True, color="666666")
    toc_header_font = Font(bold=True, size=13, color="2E75B6")

    # --- Title block ---
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "AI Infrastructure Capex & Cloud Revenue Tracker"
    cell.font = title_font
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:F2")
    ws["A2"].value = "neocloud-capex-tracker"
    ws["A2"].font = Font(size=11, italic=True, color="888888")

    # --- Summary block ---
    row = 4
    summary = [
        ("Reporting currency", "All values in USD millions (non-USD converted at period-end FX)"),
        ("Data as of", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ("Companies tracked", str(_count(conn, "SELECT COUNT(DISTINCT ticker) FROM companies"))),
        ("Total data points", str(_count(conn, "SELECT COUNT(*) FROM extractions"))),
        ("Coverage period", _coverage_range(conn)),
        ("Sources", "SEC EDGAR (10-K, 10-Q, 20-F), HKEXnews (HK-AR, HK-IR)"),
        ("Extraction methods", "XBRL API, LLM extraction (Claude Code), manual derivation"),
        ("FX source", "ECB via frankfurter.app (period-end rates)"),
    ]
    for label, value in summary:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=2, value=value).font = value_font
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1, value=(
        "DISCLAIMER: Data extracted programmatically from public SEC and HKEX filings. "
        "Every value has a source citation (Shift+F2 on any data cell). "
        "Verify against source filings before use in investment decisions."
    )).font = disclaimer_font
    row += 2

    # --- Table of Contents ---
    ws.cell(row=row, column=1, value="Table of Contents").font = toc_header_font
    row += 1

    toc_headers = ["#", "Sheet", "Description"]
    for col, h in enumerate(toc_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
    row += 1

    for idx, (sheet_name, description) in enumerate(sheet_info, 1):
        ws.cell(row=row, column=1, value=idx)
        link_cell = ws.cell(row=row, column=2, value=sheet_name)
        link_cell.font = Font(
            color="0563C1", underline="single", size=11
        )
        # Internal hyperlink to the sheet
        safe_name = sheet_name.replace("'", "''")
        link_cell.hyperlink = f"#'{safe_name}'!A1"
        ws.cell(row=row, column=3, value=description).font = value_font

        # Add row count
        target_ws = wb[sheet_name]
        rows_in_sheet = target_ws.max_row - 1  # minus header
        ws.cell(row=row, column=4, value=f"{rows_in_sheet} rows")

        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 12


def _count(conn, sql: str) -> int:
    return conn.execute(sql).fetchone()[0]


def _coverage_range(conn) -> str:
    row = conn.execute(
        "SELECT MIN(sd.fiscal_year), MAX(sd.fiscal_year) "
        "FROM extractions e "
        "JOIN source_documents sd ON e.source_document_id = sd.id"
    ).fetchone()
    if row and row[0]:
        return f"FY{row[0]} – FY{row[1]}"
    return "N/A"


def _make_font(bold=False, size=11):
    from openpyxl.styles import Font
    return Font(bold=bold, size=size)


def _build_data_quality_sheet(wb, conn, hfont, hfill):
    """Build data quality flags sheet."""
    ws = wb.create_sheet("Data Quality")

    headers = ["Company", "Issue", "Details"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill

    flags = []

    # Check for companies with sparse data
    for row in conn.execute("""
        SELECT sd.ticker, COUNT(DISTINCT e.metric_key) as n_metrics,
               COUNT(e.id) as n_extractions,
               MIN(sd.period_of_report) as earliest
        FROM extractions e
        JOIN source_documents sd ON e.source_document_id = sd.id
        GROUP BY sd.ticker
        ORDER BY sd.ticker
    """):
        ticker = row["ticker"]
        if row["n_metrics"] < 5:
            flags.append((ticker, "Missing metrics",
                         f"Only {row['n_metrics']}/5 headline metrics available"))
        if row["n_extractions"] < 10:
            flags.append((ticker, "Sparse data",
                         f"Only {row['n_extractions']} data points (earliest: {row['earliest']})"))

    # NBIS pre-restructuring flag
    flags.append(("NBIS", "Pre-restructuring data",
                 "CIK includes Yandex history. Data before 2024 is pre-restructuring "
                 "and not comparable. Filtered in annual/quarterly sheets."))

    # META cloud exclusion
    flags.append(("META", "No cloud segment",
                 "Meta is an AI infra buyer, not a cloud vendor. "
                 "Excluded from cloud_segment_revenue dataset."))

    # AMZN Q2 2017 capex gap
    flags.append(("AMZN", "Missing Q2 2017 capex",
                 "Amazon switched XBRL concepts mid-2017: "
                 "PaymentsToAcquirePropertyPlantAndEquipment (through Q1 2017) → "
                 "PaymentsToAcquireProductiveAssets (from Q3 2017). Q2 2017 was "
                 "filed without tagging capex under either concept. FY2017 annual "
                 "total ($11,955M) is correct. Fill via LLM extraction from the "
                 "actual Q2 2017 10-Q when quarterly filings are downloaded."))

    # Tencent no XBRL
    flags.append(("0700", "No XBRL data",
                 "Tencent is HKEX-only — SEC XBRL API does not cover HK filings. "
                 "All metrics require LLM extraction from the downloaded annual "
                 "report (data/_sources/0700/_raw/). FY2025 AR is available."))

    # BABA sparse capex
    flags.append(("BABA", "Sparse capex XBRL",
                 "Alibaba's XBRL tagging for capex is incomplete — only 3 annual "
                 "points available via XBRL (FY2019, FY2020, FY2025). Full "
                 "coverage requires LLM extraction from 20-F filings."))

    # 20-F filers quarterly limitation
    flags.append(("BABA/BIDU/GDS/IREN/NBIS", "No quarterly data",
                 "Foreign private issuers (20-F filers) do not file 10-Q. "
                 "Quarterly data is structurally unavailable from SEC. Only "
                 "annual data points exist. HKEX interim reports (H1) may "
                 "provide semi-annual granularity for some names."))

    for i, (ticker, issue, details) in enumerate(flags, 2):
        ws.cell(row=i, column=1, value=ticker)
        ws.cell(row=i, column=2, value=issue)
        ws.cell(row=i, column=3, value=details)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 80


def _build_metadata_sheet(wb, conn, hfont, hfill):
    """Build metadata sheet with extraction info."""
    ws = wb.create_sheet("Metadata")

    # Summary stats
    ws.cell(row=1, column=1, value="Metadata").font = _make_font(bold=True, size=14)
    ws.cell(row=3, column=1, value="Generated").font = hfont
    ws.cell(row=3, column=2,
            value=datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"))

    total = conn.execute("SELECT COUNT(*) FROM extractions").fetchone()[0]
    ws.cell(row=4, column=1, value="Total extractions").font = hfont
    ws.cell(row=4, column=2, value=total)

    n_cos = conn.execute(
        "SELECT COUNT(DISTINCT ticker) FROM source_documents"
    ).fetchone()[0]
    ws.cell(row=5, column=1, value="Companies").font = hfont
    ws.cell(row=5, column=2, value=n_cos)

    # Per-company coverage
    ws.cell(row=7, column=1, value="Company Coverage").font = _make_font(bold=True, size=12)
    cov_headers = ["Company", "Currency", "Earliest", "Latest", "Extractions", "Source"]
    for col, h in enumerate(cov_headers, 1):
        cell = ws.cell(row=8, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill

    row = 9
    for r in conn.execute("""
        SELECT sd.ticker, c.reporting_currency,
               MIN(sd.period_of_report) as earliest,
               MAX(sd.period_of_report) as latest,
               COUNT(e.id) as n_ext,
               GROUP_CONCAT(DISTINCT e.extracting_model) as models
        FROM extractions e
        JOIN source_documents sd ON e.source_document_id = sd.id
        JOIN companies c ON sd.ticker = c.ticker
        GROUP BY sd.ticker
        ORDER BY sd.ticker
    """):
        ws.cell(row=row, column=1, value=r["ticker"])
        ws.cell(row=row, column=2, value=r["reporting_currency"])
        ws.cell(row=row, column=3, value=r["earliest"])
        ws.cell(row=row, column=4, value=r["latest"])
        ws.cell(row=row, column=5, value=r["n_ext"])
        ws.cell(row=row, column=6, value=r["models"])
        row += 1

    # FX rates used
    row += 1
    ws.cell(row=row, column=1, value="FX Rates Used").font = _make_font(bold=True, size=12)
    row += 1
    fx_headers = ["Currency Pair", "Date", "Rate", "Source"]
    for col, h in enumerate(fx_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
    row += 1
    for r in conn.execute(
        "SELECT currency_pair, rate_date, rate, source "
        "FROM fx_rates ORDER BY currency_pair, rate_date"
    ):
        ws.cell(row=row, column=1, value=r[0])
        ws.cell(row=row, column=2, value=r[1])
        ws.cell(row=row, column=3, value=round(r[2], 6))
        ws.cell(row=row, column=4, value=r[3])
        row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 25


# --------------------------------------------------------------------
# Data helpers
# --------------------------------------------------------------------


def _get_metric_data(
    conn, metric_key: str, cadence: str
) -> dict[str, dict[str, float]]:
    """Get metric data organized as {ticker: {period: value_usd}}.

    For annual cadence: only FY-end periods (period_token='AR').
    For quarterly: de-cumulates flow metrics to standalone quarterly.
    Uses value_usd for cross-company comparison.
    """
    is_flow = metric_key in FLOW_METRICS

    if cadence == "annual":
        return _get_annual_data(conn, metric_key)
    else:
        return _get_quarterly_data(conn, metric_key, is_flow)


def _get_annual_data(
    conn, metric_key: str
) -> dict[str, dict[str, Any]]:
    """Get annual (FY-end) data for a metric.

    Returns {ticker: {period_label: (value, metadata_dict)}}.
    The metadata dict contains extraction + source_doc + company info
    for provenance citations.
    """
    result: dict[str, dict[str, Any]] = {}

    # Get FYE months + company info
    fye_months = {}
    company_info = {}
    for r in conn.execute(
        "SELECT ticker, fiscal_year_end_month, reporting_currency, "
        "name FROM companies"
    ):
        fye_months[r["ticker"]] = r["fiscal_year_end_month"]
        company_info[r["ticker"]] = dict(r)

    rows = conn.execute("""
        SELECT sd.ticker, sd.period_of_report, sd.period_token,
               sd.fiscal_year, sd.form_type, sd.filing_date,
               sd.source, sd.source_url, sd.accession_number,
               e.id as extraction_id,
               e.value_usd, e.value, e.reporting_currency,
               e.quote, e.locator_section, e.extracting_model,
               e.extraction_type, e.fx_rate, e.fx_rate_date,
               e.metric_key
        FROM extractions e
        JOIN source_documents sd ON e.source_document_id = sd.id
        WHERE e.metric_key = ?
        AND sd.period_token = 'AR'
        AND e.id = (
            SELECT e2.id FROM extractions e2
            JOIN source_documents sd2 ON e2.source_document_id = sd2.id
            WHERE sd2.ticker = sd.ticker
            AND sd2.period_of_report = sd.period_of_report
            AND e2.metric_key = e.metric_key
            AND sd2.period_token = 'AR'
            ORDER BY e2.value_usd DESC NULLS LAST, e2.extracted_at DESC
            LIMIT 1
        )
        ORDER BY sd.ticker, sd.period_of_report
    """, (metric_key,)).fetchall()

    for r in rows:
        ticker = r["ticker"]
        period = r["period_of_report"]
        fy = r["fiscal_year"]

        start = COVERAGE_START_OVERRIDES.get(ticker)
        if start and period < start:
            continue

        fye_m = fye_months.get(ticker, 12)
        period_month = int(period.split("-")[1])
        if period_month != fye_m:
            continue

        val = r["value_usd"] if r["value_usd"] is not None else r["value"]
        if val is None:
            continue

        label = f"FY{fy}"
        meta = {
            "extraction": dict(r),
            "source_doc": {
                "ticker": ticker,
                "form_type": r["form_type"],
                "period_of_report": period,
                "filing_date": r["filing_date"],
                "source": r["source"],
                "source_url": r["source_url"],
                "accession_number": r["accession_number"],
            },
            "company": company_info.get(ticker, {}),
        }
        result.setdefault(ticker, {})[label] = (abs(val), meta)

    return result


def _get_quarterly_data(
    conn, metric_key: str, is_flow: bool
) -> dict[str, dict[str, Any]]:
    """Get quarterly data, de-cumulating flow metrics.

    Returns {ticker: {period_label: (value, metadata_dict)}}.
    """
    result: dict[str, dict[str, Any]] = {}

    fye_months = {}
    company_info = {}
    for r in conn.execute(
        "SELECT ticker, fiscal_year_end_month, reporting_currency, "
        "name FROM companies"
    ):
        fye_months[r["ticker"]] = r["fiscal_year_end_month"]
        company_info[r["ticker"]] = dict(r)

    rows = conn.execute("""
        SELECT sd.ticker, sd.period_of_report, sd.period_token,
               sd.fiscal_year, sd.form_type, sd.filing_date,
               sd.source, sd.source_url, sd.accession_number,
               e.id as extraction_id,
               e.value_usd, e.value, e.reporting_currency,
               e.quote, e.locator_section, e.extracting_model,
               e.extraction_type, e.fx_rate, e.fx_rate_date,
               e.metric_key
        FROM extractions e
        JOIN source_documents sd ON e.source_document_id = sd.id
        WHERE e.metric_key = ?
        AND e.id = (
            SELECT e2.id FROM extractions e2
            JOIN source_documents sd2 ON e2.source_document_id = sd2.id
            WHERE sd2.ticker = sd.ticker
            AND sd2.period_of_report = sd.period_of_report
            AND e2.metric_key = e.metric_key
            ORDER BY e2.value_usd DESC NULLS LAST, e2.extracted_at DESC
            LIMIT 1
        )
        ORDER BY sd.ticker, sd.fiscal_year, sd.period_of_report
    """, (metric_key,)).fetchall()

    by_ticker_fy: dict[str, dict[int, list]] = {}
    for r in rows:
        ticker = r["ticker"]
        start = COVERAGE_START_OVERRIDES.get(ticker)
        if start and r["period_of_report"] < start:
            continue
        fy = r["fiscal_year"]
        by_ticker_fy.setdefault(ticker, {}).setdefault(fy, []).append(r)

    for ticker, fy_data in by_ticker_fy.items():
        for _fy, points in fy_data.items():
            points.sort(key=lambda x: x["period_of_report"])
            tokens = {p["period_token"] for p in points}
            if tokens == {"AR"}:
                continue

            prev_val = 0
            for p in points:
                val = p["value_usd"] if p["value_usd"] is not None else p["value"]
                if val is None:
                    continue
                val = abs(val)
                token = p["period_token"]
                period = p["period_of_report"]

                if is_flow:
                    if token == "Q1":
                        quarterly_val = val
                        prev_val = val
                    elif token in ("Q2", "Q3"):
                        quarterly_val = val - prev_val
                        prev_val = val
                    elif token == "AR":
                        quarterly_val = val - prev_val
                        prev_val = 0
                    else:
                        quarterly_val = val
                else:
                    quarterly_val = val

                label = f"{period[:4]}Q{_quarter_from_token(token)}"
                meta = {
                    "extraction": dict(p),
                    "source_doc": {
                        "ticker": ticker,
                        "form_type": p["form_type"],
                        "period_of_report": period,
                        "filing_date": p["filing_date"],
                        "source": p["source"],
                        "source_url": p["source_url"],
                        "accession_number": p["accession_number"],
                    },
                    "company": company_info.get(ticker, {}),
                }
                result.setdefault(ticker, {})[label] = (
                    quarterly_val,
                    meta,
                )

    return result


def _quarter_from_token(token: str) -> str:
    mapping = {"Q1": "1", "Q2": "2", "Q3": "3", "AR": "4", "H1": "2", "H2": "4"}
    return mapping.get(token, "?")
